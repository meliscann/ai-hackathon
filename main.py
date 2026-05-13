"""
KOBİ Pilot — FastAPI backend
Çalıştır: uvicorn main:app --reload --port 8000
"""

import sqlite3
import json
from datetime import datetime
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import init_db, seed_data
from agent import KobiAgent

# ── App ─────────────────────────────────────────────────────────────────────
app = FastAPI(title="KOBİ Pilot API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

agent = KobiAgent()


# ── WebSocket Connection Manager ─────────────────────────────────────────────
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        dead = []
        for ws in self.active_connections:
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


manager = ConnectionManager()


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)


# ── Startup ───────────────────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    init_db()
    seed_data()


# ── Schemas ──────────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    message: str
    session_id: str = "default"


class ChatResponse(BaseModel):
    response: str


class StatusUpdate(BaseModel):
    status: str


class StockUpdate(BaseModel):
    stock: int


class ProductCreate(BaseModel):
    name: str
    category: str
    price: float
    stock: int
    min_stock: int
    unit: str = "adet"


class ActionRequest(BaseModel):
    product_name: str
    quantity: int = 50


# ── Chat ─────────────────────────────────────────────────────────────────────
@app.post("/chat", response_model=ChatResponse)
def chat(req: ChatRequest):
    try:
        reply = agent.chat(req.session_id, req.message)
        return ChatResponse(response=reply)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/reset/{session_id}")
def reset_session(session_id: str):
    agent.reset(session_id)
    return {"status": "ok"}


@app.post("/generate-action")
def generate_action(req: ActionRequest):
    try:
        reply = agent.generate_action(req.product_name, req.quantity)
        return {"action_text": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Dashboard ────────────────────────────────────────────────────────────────
@app.get("/dashboard")
def dashboard():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")

    c.execute(
        "SELECT COUNT(*), COALESCE(SUM(total_price), 0) FROM orders WHERE created_at LIKE ?",
        (f"{today}%",),
    )
    today_cnt, today_rev = c.fetchone()

    c.execute("SELECT COUNT(*) FROM orders WHERE status='beklemede'")
    pending = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM products WHERE stock <= min_stock")
    low_stock = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM orders WHERE status='kargoda'")
    in_cargo = c.fetchone()[0]

    c.execute("SELECT status, COUNT(*) FROM orders GROUP BY status")
    status_dist = {r[0]: r[1] for r in c.fetchall()}

    conn.close()

    return {
        "today_orders":        today_cnt,
        "today_revenue":       round(today_rev, 2),
        "pending_orders":      pending,
        "low_stock_count":     low_stock,
        "in_cargo":            in_cargo,
        "status_distribution": status_dist,
    }


# ── Orders ───────────────────────────────────────────────────────────────────
@app.get("/orders")
def get_orders(status: str = None, limit: int = 50):
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()

    sql = """
        SELECT o.id, o.customer_name, p.name, o.quantity, o.total_price,
               o.status, o.cargo_tracking, o.created_at, o.estimated_delivery, o.lat, o.lng
        FROM orders o
        JOIN products p ON o.product_id = p.id
        {where}
        ORDER BY o.created_at DESC LIMIT ?
    """
    if status:
        c.execute(sql.format(where="WHERE o.status=?"), (status, limit))
    else:
        c.execute(sql.format(where=""), (limit,))

    rows = c.fetchall()
    conn.close()

    return [
        {
            "id":       r[0], "customer": r[1], "product": r[2],
            "quantity": r[3], "total":    round(r[4], 2), "status": r[5],
            "tracking": r[6], "created":  r[7], "delivery": r[8],
            "lat":      r[9], "lng":      r[10],
        }
        for r in rows
    ]


@app.patch("/orders/{order_id}/status")
async def update_order_status(order_id: int, body: StatusUpdate):
    valid = {"beklemede", "hazırlanıyor", "kargoda", "teslim edildi", "iptal"}
    if body.status not in valid:
        raise HTTPException(400, f"Geçersiz durum. Geçerli değerler: {sorted(valid)}")

    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute("UPDATE orders SET status=? WHERE id=?", (body.status, order_id))
    if c.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Sipariş bulunamadı")
    conn.commit()
    conn.close()

    await manager.broadcast({
        "type":    "order_update",
        "order_id": order_id,
        "status":  body.status,
        "message": f"Sipariş #{order_id} → {body.status}",
    })
    return {"status": "ok"}


# ── Products ─────────────────────────────────────────────────────────────────
@app.get("/products")
def get_products():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute(
        "SELECT id, name, category, price, stock, min_stock, unit FROM products ORDER BY stock ASC"
    )
    rows = c.fetchall()
    conn.close()

    return [
        {
            "id": r[0], "name": r[1], "category": r[2],
            "price": r[3], "stock": r[4], "min_stock": r[5],
            "unit": r[6], "low": r[4] <= r[5],
        }
        for r in rows
    ]


@app.post("/products", status_code=201)
async def add_product(product: ProductCreate):
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute(
        "INSERT INTO products (name, category, price, stock, min_stock, unit) VALUES (?,?,?,?,?,?)",
        (product.name, product.category, product.price, product.stock, product.min_stock, product.unit),
    )
    new_id = c.lastrowid
    conn.commit()
    conn.close()

    await manager.broadcast({
        "type":    "product_added",
        "message": f"Yeni ürün eklendi: {product.name}",
    })
    return {"id": new_id, **product.model_dump()}


@app.delete("/products/{product_id}", status_code=204)
async def delete_product(product_id: int):
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM orders WHERE product_id=?", (product_id,))
    if c.fetchone()[0] > 0:
        conn.close()
        raise HTTPException(400, "Bu ürüne ait siparişler bulunduğundan silinemez")
    c.execute("DELETE FROM products WHERE id=?", (product_id,))
    if c.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Ürün bulunamadı")
    conn.commit()
    conn.close()
    await manager.broadcast({"type": "product_deleted", "message": "Ürün silindi"})


@app.patch("/products/{product_id}/stock")
async def update_stock(product_id: int, body: StockUpdate):
    if body.stock < 0:
        raise HTTPException(400, "Stok miktarı negatif olamaz")

    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute("UPDATE products SET stock=? WHERE id=?", (body.stock, product_id))
    if c.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Ürün bulunamadı")
    c.execute("SELECT name, min_stock FROM products WHERE id=?", (product_id,))
    row = c.fetchone()
    conn.commit()
    conn.close()

    if row and body.stock <= row[1]:
        await manager.broadcast({
            "type":         "low_stock",
            "product_id":   product_id,
            "product_name": row[0],
            "stock":        body.stock,
            "message":      f"Kritik Stok: {row[0]} ({body.stock} adet kaldı)",
        })
    return {"status": "ok"}


# ── Analytics & Forecast ──────────────────────────────────────────────────────
@app.get("/analytics")
def get_analytics():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()

    c.execute("SELECT status, COUNT(*) FROM orders GROUP BY status")
    status_dist = {r[0]: r[1] for r in c.fetchall()}

    c.execute("""
        SELECT date(created_at) as d, COUNT(*) as cnt
        FROM orders
        GROUP BY d
        ORDER BY d DESC LIMIT 7
    """)
    daily_traffic = [{"date": r[0], "orders": r[1]} for r in c.fetchall()]
    daily_traffic.reverse()

    conn.close()
    return {"status_distribution": status_dist, "daily_traffic": daily_traffic}


@app.get("/forecast")
def get_forecast():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM orders WHERE created_at >= date('now', '-7 days')")
    last_7_days = c.fetchone()[0]
    conn.close()

    forecast_orders = int(last_7_days * 1.15)
    return {
        "forecast_message": f"Gelecek hafta {forecast_orders} sipariş bekleniyor.",
        "forecast_orders":  forecast_orders,
        "growth":           "15%",
    }


# ── Customers ─────────────────────────────────────────────────────────────────
@app.get("/customers")
def get_customers():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()
    c.execute("""
        SELECT
            o.customer_name,
            o.customer_phone,
            COUNT(o.id)                         AS order_count,
            COALESCE(SUM(o.total_price), 0)     AS total_spent,
            MAX(o.created_at)                   AS last_order,
            GROUP_CONCAT(DISTINCT p.name)       AS products
        FROM orders o
        JOIN products p ON o.product_id = p.id
        GROUP BY o.customer_name, o.customer_phone
        ORDER BY total_spent DESC
    """)
    rows = c.fetchall()
    conn.close()

    result = []
    for r in rows:
        total = round(r[3], 2)
        count = r[2]
        segment = "VIP" if total >= 500 else ("Sadık" if count >= 3 else "Yeni")
        result.append({
            "name":        r[0],
            "phone":       r[1],
            "order_count": count,
            "total_spent": total,
            "last_order":  r[4],
            "products":    r[5].split(",") if r[5] else [],
            "segment":     segment,
        })
    return result


# ── Excel Report ──────────────────────────────────────────────────────────────
@app.get("/report/excel")
def download_excel():
    conn = sqlite3.connect("kobi_pilot.db")
    c = conn.cursor()

    wb = openpyxl.Workbook()

    # ── Siparişler sayfası ──
    ws = wb.active
    ws.title = "Siparişler"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D9488")
    headers = ["Sipariş No", "Müşteri", "Telefon", "Ürün", "Adet", "Tutar (TL)", "Durum", "Kargo Takip", "Tarih", "Tahmini Teslimat"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    c.execute("""
        SELECT o.id, o.customer_name, o.customer_phone, p.name,
               o.quantity, o.total_price, o.status,
               COALESCE(o.cargo_tracking, '-'), o.created_at, o.estimated_delivery
        FROM orders o
        JOIN products p ON o.product_id = p.id
        ORDER BY o.created_at DESC
    """)
    for row in c.fetchall():
        ws.append(list(row))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    # ── Stok Durumu sayfası ──
    ws2 = wb.create_sheet("Stok Durumu")
    ws2.append(["Ürün", "Kategori", "Fiyat (TL)", "Stok", "Min Stok", "Birim", "Durum"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    c.execute("SELECT name, category, price, stock, min_stock, unit FROM products ORDER BY stock ASC")
    for row in c.fetchall():
        durum = "KRİTİK" if row[3] <= row[4] else "Normal"
        ws2.append(list(row) + [durum])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 16

    # ── Müşteri Özeti sayfası ──
    ws3 = wb.create_sheet("Müşteriler")
    ws3.append(["Ad Soyad", "Telefon", "Sipariş Sayısı", "Toplam Harcama (TL)", "Son Sipariş", "Segment"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    c.execute("""
        SELECT o.customer_name, o.customer_phone,
               COUNT(o.id), COALESCE(SUM(o.total_price), 0), MAX(o.created_at)
        FROM orders o
        GROUP BY o.customer_name, o.customer_phone
        ORDER BY SUM(o.total_price) DESC
    """)
    for row in c.fetchall():
        total = round(row[3], 2)
        segment = "VIP" if total >= 500 else ("Sadık" if row[2] >= 3 else "Yeni")
        ws3.append([row[0], row[1], row[2], total, row[4], segment])
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"kobi_pilot_rapor_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
